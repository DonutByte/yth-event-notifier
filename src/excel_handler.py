from openpyxl import load_workbook
import openpyxl.cell
from openpyxl.utils.cell import column_index_from_string
import datetime
from event import Event
import requests
import re
import os

# from creds import EXCEL_URL
EXCEL_URL = os.environ['EXCEL_URL']
DOWNLOAD_URL = f'{EXCEL_URL}/export?format=xlsx'


class ExcelWorker:
    VALID_GRADE_COLUMNS = re.compile(r"([טיאב']+)\d")
    GRADES_ROW = 2
    GRADES = {9: (None, None), 10: (None, None),
              11: (None, None), 12: (None, None)}
    BAGROT_COLOR = ['FF0000FF']
    TEST_COLOR = ['00000000']
    INSIDE_BAGROT_COLOR = ['FF3D85C6', 'FF6D9EEB']
    DATE_COLUMN = 'E'

    def __init__(self, workbook_path: str, intervals):
        if not os.path.exists(workbook_path):
            file = requests.get(DOWNLOAD_URL)
            with open(workbook_path, 'wb') as excel_file:
                excel_file.write(file.content)

        self.workbook_path = workbook_path
        self.workbook = load_workbook(self.workbook_path, read_only=True)
        years = sorted(name for name in self.workbook.sheetnames if name.startswith('תש'))[0]
        self.worksheet = self.workbook[years]
        self.schedule: dict[int, list[list[Event]]] = dict()
        self.expire_date: datetime.date = datetime.date.today()

        self.set_grades_columns()
        self.update_schedule(intervals)

    def open_worksheet(self):
        self.workbook = load_workbook(self.workbook_path, read_only=True)
        self.worksheet = self.workbook.active

    def set_grades_columns(self):
        """
        finds the columns of each grade
        :return: None, assigns the result to `self.GRADES`
        """
        grades = {9: [], 10: [], 11: [], 12: []}

        for row in self.worksheet.iter_rows(min_row=self.GRADES_ROW, max_row=self.GRADES_ROW):
            for column in row:
                if column.value is None:
                    continue
                match = self.VALID_GRADE_COLUMNS.findall(column.value)
                if match:
                    grade = self.grade_to_number(match[0])
                    grades[grade].append(column.column)

        for grade in grades:
            columns = grades[grade]
            grades[grade] = [min(columns), max(columns)]

        self.GRADES = grades

    def get_this_week_row(self):
        today = datetime.date.today()
        sunday = today - \
                 datetime.timedelta((today.weekday() + 1) % 7)  # 6 = sunday
        for cells in self.worksheet.iter_rows(max_col=column_index_from_string(self.DATE_COLUMN),
                                              min_col=column_index_from_string(
                                                  self.DATE_COLUMN),
                                              min_row=3):
            cell: openpyxl.cell.cell.Cell = cells[0]
            if sunday == cell.value.date():
                return cell.row

    def parse_today_events(self, row: tuple) -> list:
        events = []
        for event in row:
            if event and event.value and event not in events:
                events.append(event)
        return events

    def get_week_events(self, starting_col, grade) -> list[Event]:
        events: list[Event] = []
        for cells in self.worksheet.iter_rows(min_row=starting_col, max_row=starting_col + 5,
                                              min_col=self.GRADES[grade][0],
                                              max_col=self.GRADES[grade][1]):

            for cell in self.parse_today_events(cells):
                if not cell.value:  # cell is empty
                    continue
                date = self.worksheet[f'{self.DATE_COLUMN}{cell.row}'].value.date(
                )
                if cell.value.startswith('מתכ.'):
                    events.append(
                        Event(cell.value[4:].strip(" "), "מתכונת", date))
                elif cell.fill.start_color.index in self.BAGROT_COLOR:
                    events.append(Event(cell.value.strip(" "), "בגרות", date))
                elif cell.fill.start_color.index in self.INSIDE_BAGROT_COLOR:
                    events.append(
                        Event(cell.value.strip(" "), "בגרות פנימית", date))
                elif cell.fill.start_color.index in self.TEST_COLOR:
                    events.append(
                        Event(cell.value.strip(" "), "מבחן", date))
                else:
                    events.append(Event(cell.value.strip(" "), "", date))

        return events

    def update_schedule(self, intervals: list[int]):
        # get the most up-to-date version of the excel
        self.workbook.close()
        file = requests.get(DOWNLOAD_URL)
        with open(self.workbook_path, 'wb') as excel_file:
            excel_file.write(file.content)

        self.open_worksheet()  # refresh the worksheet

        row = self.get_this_week_row()
        for grade in self.GRADES:
            self.schedule[grade] = []
        for grade in self.GRADES:
            for week_interval in intervals:
                self.schedule[grade].append(
                    self.get_week_events(row + week_interval, grade))

        self.expire_date = datetime.date.today() \
                           + datetime.timedelta(days=intervals[0] * 7)

    def get_schedule(self, intervals: list[int]) -> dict[int, list[list[Event]]]:
        # before updating check if schedule's last date
        # hasn't expired
        if self.expire_date >= datetime.date.today():
            return self.schedule
        print('Getting new schedule')
        self.update_schedule(intervals)
        return self.schedule

    @staticmethod
    def grade_to_number(grade: str) -> int:
        n = 0
        for letter in grade:
            n += ord(letter) - ord('א') + 1
        return n
